"""XLSX of 24-price vectors for the 388 SC ∩ TX ∩ LIVE overlap games.

Three rows per game (SC, TX, LIVE). For each row, the columns
x_0..x_23 where THAT source's strategy fired are colored with a
source-specific fill:
  * LIVE fires (matched or errored) → orange
  * SC backtest fires               → blue
  * TX backtest fires               → green
Cells where the three sources disagree on the price are bolded
(with yellow fill when no fire color overrides).
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import numpy as np
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from pregame_dc import paths
from pregame_dc.constants import X_COLS, TYPE_FOR_SLOT, MARKET_LABELS
from pregame_dc.pipelines.helpers import canonical_slot_market_ids
from pregame_dc.analyze.full_24vec_compare import (
    parse_live_24_vectors, _detect_bot_tz_offset_ms,
)
from pregame_dc.analyze.backtest_full_overlap import (
    fire_offset_for, enabled_slots_for, load_model, build_tx_index,
    vector_24, derive_slot_mids, parse_live_game_starts,
    MODEL_BY_OFFSET, THRESHOLD, ASK_LO, ASK_HI,
)

LIVE_RESOLVED = paths.PACKAGE_ROOT / "pregame_dc" / "live" / "logs" / "orders_summary_resolved.jsonl"
BOT_STDOUT = paths.PACKAGE_ROOT / "pregame_dc" / "live" / "logs" / "bot_stdout.log"
SC_DIR = paths.PACKAGE_ROOT / "data" / "self_collected" / "per_game_data"
OUT_XLSX = paths.PACKAGE_ROOT / "plots" / "full_overlap_24vec.xlsx"


def main():
    # ---- universe ----
    sc_slugs = {p.stem for p in SC_DIR.glob("*.parquet")}
    tx_index = build_tx_index()
    live_starts = parse_live_game_starts()
    overlap = sorted(sc_slugs & set(tx_index) & set(live_starts))
    print(f"3-way overlap: {len(overlap)} games")

    # Live attempts (any status — matched + errored both count as a "fire")
    attempts = [json.loads(l) for l in LIVE_RESOLVED.open() if l.strip()]
    live_fires = defaultdict(set)   # slug -> set of slot_24 the bot attempted
    for r in attempts:
        live_fires[r["game_slug"]].add(r["slot"])

    # Parse live 24-vectors from bot_stdout
    tz_off = _detect_bot_tz_offset_ms(BOT_STDOUT, [r for r in attempts if r["status"] == "matched"])
    live_vecs = parse_live_24_vectors(BOT_STDOUT, tz_offset_ms=tz_off)

    models = {off: load_model(p) for off, p in MODEL_BY_OFFSET.items()}

    # ---- build per-game data ----
    games = []   # list of dicts: slug, sc_x, tx_x, live_x, sc_fires, tx_fires, live_fires_set
    n_skipped = 0
    for slug in overlap:
        sc_df = pq.read_table(SC_DIR / f"{slug}.parquet").to_pandas()
        tx_df = pq.read_table(tx_index[slug]).to_pandas()
        sc_mids = derive_slot_mids(sc_df)
        tx_mids = derive_slot_mids(tx_df)
        if sc_mids is None or tx_mids is None:
            n_skipped += 1
            continue
        gst = live_starts[slug]
        fire_off = fire_offset_for(gst)
        fire_time_ms = (gst + fire_off) * 1000
        enabled = set(enabled_slots_for(gst))
        sc_x = vector_24(sc_df, sc_mids, fire_time_ms)
        tx_x = vector_24(tx_df, tx_mids, fire_time_ms)
        sc_pred = models[fire_off](sc_x)
        tx_pred = models[fire_off](tx_x)
        sc_fires = set()
        tx_fires = set()
        for slot_24 in enabled:
            if ASK_LO <= sc_x[slot_24] <= ASK_HI and sc_pred[slot_24] - sc_x[slot_24] > THRESHOLD:
                sc_fires.add(slot_24)
            if ASK_LO <= tx_x[slot_24] <= ASK_HI and tx_pred[slot_24] - tx_x[slot_24] > THRESHOLD:
                tx_fires.add(slot_24)
        lv = live_vecs.get(slug)
        live_x = lv["x"] if lv else None
        live_p = lv["pred12"] if lv else None
        games.append({
            "slug": slug, "fire_offset_s": fire_off, "game_start_ts": gst,
            "sc_x": sc_x, "tx_x": tx_x, "live_x": live_x,
            "sc_pred": models[fire_off](sc_x)[:12],
            "tx_pred": models[fire_off](tx_x)[:12],
            "live_pred": live_p,
            "sc_fires": sc_fires, "tx_fires": tx_fires,
            "live_fires": live_fires.get(slug, set()),
        })
    print(f"  emitted: {len(games)} games (skipped {n_skipped} with no canonical 12-slot set)")

    # ---- build XLSX ----
    cols = ["game_slug", "fire_offset_s", "source"] \
        + X_COLS \
        + [f"pred_{i}" for i in range(12)]

    wb = Workbook()
    ws = wb.active
    ws.title = "full overlap 24vec"
    for c, h in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)

    bold = Font(bold=True)
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fire_fill = {
        "SC":   PatternFill(start_color="9FC5E8", end_color="9FC5E8", fill_type="solid"),   # blue
        "TX":   PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid"),   # green
        "LIVE": PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid"),   # orange
    }

    x_col_idx = {f"x_{i}": cols.index(f"x_{i}") + 1 for i in range(24)}
    out_row = 2

    n_total_xtrips = n_mismatched = 0
    n_fires_by_src = {"SC": 0, "TX": 0, "LIVE": 0}

    for g in games:
        sc_x, tx_x, lv_x = g["sc_x"], g["tx_x"], g["live_x"]
        sc_p, tx_p, lv_p = g["sc_pred"], g["tx_pred"], g["live_pred"]

        # Per-column disagreement check (compare to 4 decimals)
        x_disagree = {}
        for i in range(24):
            triple = []
            for v in (sc_x[i], tx_x[i], None if lv_x is None else float(lv_x[i])):
                triple.append(None if v is None else round(float(v), 4))
            x_disagree[i] = len(set(triple)) > 1
            n_total_xtrips += 1
            if x_disagree[i]:
                n_mismatched += 1

        pred_disagree = {}
        for i in range(12):
            triple = []
            for v in (sc_p[i], tx_p[i], None if lv_p is None else float(lv_p[i])):
                triple.append(None if v is None else round(float(v), 4))
            pred_disagree[i] = len(set(triple)) > 1

        rows_data = [
            ("SC", sc_x, sc_p, g["sc_fires"]),
            ("TX", tx_x, tx_p, g["tx_fires"]),
            ("LIVE", lv_x, lv_p, g["live_fires"]),
        ]

        for r_off, (source, x, p, fires) in enumerate(rows_data):
            row = out_row + r_off
            for c_idx, col_name in enumerate(cols, start=1):
                if col_name == "game_slug":
                    val = g["slug"]
                elif col_name == "fire_offset_s":
                    val = g["fire_offset_s"]
                elif col_name == "source":
                    val = source
                elif col_name.startswith("x_"):
                    i = int(col_name[2:])
                    val = None if x is None else float(x[i])
                elif col_name.startswith("pred_"):
                    i = int(col_name[5:])
                    val = None if p is None else float(p[i])
                else:
                    val = None
                cell = ws.cell(row=row, column=c_idx, value=val)
                if col_name.startswith(("x_", "pred_")) and isinstance(val, float):
                    cell.number_format = "0.0000"

                # Fire-cell coloring (only on x_i columns of the source's row)
                if col_name.startswith("x_"):
                    slot_24 = int(col_name[2:])
                    if slot_24 in fires:
                        cell.fill = fire_fill[source]
                        cell.font = bold
                        n_fires_by_src[source] += 1
                        continue

                # Disagreement: yellow + bold if no fire color was set
                if col_name.startswith("x_") and x_disagree[int(col_name[2:])]:
                    cell.fill = yellow
                    cell.font = bold
                elif col_name.startswith("pred_") and pred_disagree[int(col_name[5:])]:
                    cell.fill = yellow
                    cell.font = bold

        out_row += 3

    ws.freeze_panes = "D2"
    for c, h in enumerate(cols, start=1):
        width = 14 if h == "game_slug" else 9
        ws.column_dimensions[get_column_letter(c)].width = width

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"\nwrote {OUT_XLSX}")
    print(f"games × 3 rows = {len(games) * 3} rows")
    print(f"x-triples mismatched: {n_mismatched}/{n_total_xtrips} ({n_mismatched/n_total_xtrips*100:.1f}%)")
    print(f"fires by source (cells colored): "
          f"SC={n_fires_by_src['SC']}  TX={n_fires_by_src['TX']}  LIVE={n_fires_by_src['LIVE']}")
    print(f"colors: SC=blue, TX=green, LIVE=orange; disagreements=yellow")


if __name__ == "__main__":
    main()
